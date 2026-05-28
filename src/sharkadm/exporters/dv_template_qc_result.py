import openpyxl
import pathlib
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from sharkadm.exporters.base import PolarsFileExporter
from sharkadm.data.data_holder import PolarsDataHolder


class ExportDvTemplateWithQcResult(PolarsFileExporter):
    def __init__(self,
                 dv_template_file,
                 **kwargs):
        self._dv_template_file = pathlib.Path(dv_template_file)
        super().__init__(**kwargs)
        if not kwargs.get("export_file_name"):
            self._export_file_name = f"{self._dv_template_file.stem}_qc_result.xlsx"

    @staticmethod
    def get_exporter_description() -> str:
        return "Export qc result as DV template data"

    def _export(self, data_holder: PolarsDataHolder) -> None:
        indata = openpyxl.load_workbook(self._dv_template_file)
        tab = indata['Kolumner']


        parameter_to_column = {}
        for col in tab.iter_cols(min_row=3, max_row=3):
            for cell in col:
                if cell.value:
                    parameter_to_column[cell.value] = cell.column_letter

        print("Dynamisk mappning (från rad 3):")
        print(parameter_to_column)

        unique_parameters = sorted(
            {row["parameter"] for row in data_holder.data.iter_rows(named=True) if row["parameter"] in parameter_to_column},
            key=lambda p: openpyxl.utils.column_index_from_string(parameter_to_column[p]),
            reverse=True
        )

        for parameter in unique_parameters:
            param_column = parameter_to_column[parameter]
            param_col_index = openpyxl.utils.column_index_from_string(param_column)
            total_qc_col_index = param_col_index + 1

            tab.insert_cols(total_qc_col_index)
            total_qc_column = get_column_letter(total_qc_col_index)

            tab[f"{total_qc_column}3"] = f"TOTAL_QC_{parameter}"

            for row in data_holder.data.iter_rows(named=True):
                if row["parameter"] == parameter:
                    rad = row["row_number"]
                    excel_row = int(rad) + 3
                    cell_ref = f"{total_qc_column}{excel_row}"

                    tab[cell_ref] = row["TOTAL_QC"]

                    if row.get("total_automatic_info"):
                        cell = tab[cell_ref]
                        cell.comment = Comment(
                            text=row["total_automatic_info"],
                            author="SHARKadm QC Tool"
                        )

        all_parameters = {row["parameter"] for row in data_holder.data.iter_rows(named=True)}
        missing_parameters = all_parameters - parameter_to_column.keys()
        if missing_parameters:
            print(f"Varning: Följande parametrar hittades inte i Excel-filen: {missing_parameters}")

        indata.save(self.export_file_path)# Savea som ny fil
        print("TOTAL_QC-kolumner har lagts till och fyllts i Excel-filen.")