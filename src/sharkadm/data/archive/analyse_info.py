# -*- coding: utf-8 -*-

import datetime
import logging
import pathlib
import re
from typing import Protocol

import openpyxl
import pandas as pd

from sharkadm import adm_logger


DATE_FORMATS = [
    '%Y-%m-%d',
    '%Y-%m'
    ]

logger = logging.getLogger(__name__)


class Mapper(Protocol):

    def get_internal_name(self, external_par: str) -> str:
        ...


class UncertString:
    def __init__(self, string: str):
        self._string = string.strip()

        self._value: str = None
        self._percent: int = None
        self._interval = []

        self._save_data()

    @property
    def value(self) -> str | None:
        if self._value is None:
            return None
        return str(self._value)

    @property
    def percent(self) -> str | None:
        if not self._percent:
            return None
        return f'{self._percent} %'

    def __repr__(self):
        return self.value or self.percent

    def _save_data(self):
        if not self._string:
            return
        if self._string.lower() in ['calculated']:
            self._value = self._string.capitalize()
            return
        try:
            float(self._string)
            self._value = self._string
            return
        except:
            pass
        if '%' in self._string:
            self._percent = int(self._string.split('%')[0].strip())
        else:
            self._value = re.search('[0-9\.]+', self._string).group()

        if '-' in self._string:
            range = re.search('[0-9\.-]+', self._string.replace(' ', '').split('%')[-1]).group()
            self._interval = [float(item.strip()) for item in range.split('-')]

    def get(self, value: str | float | int):
        val = float(value)
        if self._interval and not (self._interval[0] <= val <= self._interval[-1]):
            return False
        if self._percent:
            return str(round((val * self._percent / 100), 2))
        return str(self._value)


class UncertComment:
    def __init__(self, comment: str | float | int):
        self._orig_comment = str(comment)
        self._info = []
        self._save_data()

    def _save_data(self):
        for string in self._orig_comment.split(','):
            self._info.append(UncertString(string))

    def get(self, value: str | float | int) -> str:
        for item in self._info:
            val = item.get(value)
            if val:
                return str(val)


class AnalyseInfo:

    def __init__(self, data: dict, mapper: Mapper = None) -> None:
        self._data = data
        self._path = data.pop('path', None)
        self._mapper = mapper

        if self._mapper:
            self._map_data()

    def _map_data(self):
        if not self._mapper:
            adm_logger.log_workflow(f'No mapper found when trying to map AnalyseInfo file: {self._path}')
            return
        new_data = {}
        for par, info_list in self._data.items():
            internal_par = self._mapper.get_internal_name(par)
            if internal_par.startswith('COPY_VARIABLE'):
                internal_par = internal_par.split('.')[1]
            new_data.setdefault(internal_par, [])
            new_data.setdefault(par, [])
            for info in info_list:
                new_info = dict()
                for key, value in info.items():
                    internal_key = self._mapper.get_internal_name(key)
                    new_info[internal_key] = value
                new_data[internal_par].append(new_info)
                new_data[par].append(new_info)
        self._data = new_data

    def __str__(self):
        lst = '\n'.join([sorted(self.data)])
        return f'Analyse info for parameters: \n{lst}'

    def get_info(self, par: str, date: datetime.date) -> dict:
        info = {}
        for info in self.data.get(par, []):
            if info['VALIDFR'] and info['VALIDTO'] and (info['VALIDFR'] <= date <= info['VALIDTO']):
                return info
            if info['VALIDFR'] and (info['VALIDFR'] <= date):
                return info
            if info['VALIDTO'] and (date <= info['VALIDTO']):
                return info
        return info

    def get(self, par: str, col: str, date: datetime.date) -> str | None:
        info = self.get_info(par=par, date=date)
        return info.get(col)

    def get_uncertainty(self, par: str, date: datetime.date, value: float | str | None = None) -> str:
        comment = self.get(par=par, date=date, col='UNCERT')
        if value is None:
            return comment
        obj = UncertComment(comment)
        return obj.get(value)


    @classmethod
    def from_txt_file(cls, path: str | pathlib.Path, mapper: Mapper = None, encoding: str = 'cp1252') -> 'AnalyseInfo':
        path = pathlib.Path(path)
        if path.suffix != '.txt':
            msg = f'File is not a valid analyse_info text file: {path}'
            logger.error(msg)
            raise FileNotFoundError(msg)
        data = dict()
        data['path'] = path
        with open(path, encoding=encoding) as fid:
            header = []
            for r, line in enumerate(fid):
                if not line.strip():
                    continue
                split_line = [item.strip() for item in line.split('\t')]
                if r == 0:
                    header = split_line
                    continue
                if len(split_line) != len(header):
                    adm_logger.log_workflow(f'No or insufficient data in analyse_info for parameter: {split_line[0]}', level=adm_logger.WARNING)
                    continue
                line_dict = dict(zip(header, split_line))

                for item in ['VALIDFR', 'VALIDTO']:
                    if not line_dict.get(item):
                        adm_logger.log_workflow(f'Missing {item} in analyse_info', level=adm_logger.WARNING)
                        line_dict[item] = None
                        continue
                    line_dict[item] = _get_date(line_dict[item])

                par = line_dict['PARAM']
                data.setdefault(par, [])
                data[par].append(line_dict)
        return AnalyseInfo(data, mapper=mapper)

    @classmethod
    def from_lims_txt_file(cls, path: str | pathlib.Path, mapper: Mapper, encoding: str = 'cp1252') -> 'AnalyseInfo':
        path = pathlib.Path(path)
        if path.suffix != '.txt':
            msg = f'File is not a valid analyse_info text file: {path}'
            logger.error(msg)
            raise FileNotFoundError(msg)
        data = dict()
        data['path'] = path
        with open(path, encoding=encoding) as fid:
            header = []
            for r, line in enumerate(fid):
                if not line.strip():
                    continue
                split_line = [item.strip() for item in line.split('\t')]
                if r == 0:
                    header = split_line
                    continue
                if len(split_line) != len(header):
                    adm_logger.log_workflow(f'No or insufficient data in analyse_info for parameter: {split_line[0]}',
                                            level=adm_logger.WARNING)
                    continue
                line_dict = dict(zip(header, split_line))
                line_dict['VALIDFR'] = _get_date(line_dict['VALIDFR (YYYY-MM-DD)'])
                line_dict['VALIDTO'] = _get_date(line_dict['VALIDTO'])
                par = line_dict['PARAM']
                data.setdefault(par, [])
                data[par].append(line_dict)
        return AnalyseInfo(data, mapper=mapper)

    @classmethod
    def from_dv_template(cls, path: str | pathlib.Path, mapper: Mapper = None) -> 'AnalyseInfo':
        wb = openpyxl.load_workbook(path)
        sheet_name = None
        for name in ['Analysinfo']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            adm_logger.log_workflow(f'Could not find analyse_info sheet in file: {path}', level=adm_logger.WARNING)
            return
            # raise Exception(f'Could not find analyse_info sheet in file: {path}')

        ws = wb.get_sheet_by_name(sheet_name)
        skip_nr_rows = 0
        for r in range(1, 5):
            if ws.cell(r, 1).value in ['Tabellhuvud:']:
                skip_nr_rows = r - 1
                break
        df = pd.read_excel(path, skiprows=skip_nr_rows, sheet_name=sheet_name, dtype=str)
        df.fillna('', inplace=True)
        data = dict()
        data['path'] = path
        for row in df.iterrows():
            line_dict = row[1].to_dict()
            line_dict['VALIDFR'] = _get_date(line_dict['VALIDFR'])
            line_dict['VALIDTO'] = _get_date(line_dict['VALIDTO'])
            par = line_dict['PARAM']
            data.setdefault(par, [])
            data[par].append(line_dict)
        return AnalyseInfo(data, mapper=mapper)

    @property
    def data(self) -> dict[str, str]:
        return self._data

    @property
    def parameters(self) -> list[str]:
        """Returns a list of all parameters in the file. The list is unsorted."""
        return list(self._data)

    @property
    def columns(self) -> list[str]:
        return list(self.data[list(self._data)[0]][0])


def _get_date(date_str: str) -> datetime.date | str:
    if not date_str:
        return ''
    d_string = date_str.split()[0]
    for form in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(d_string, form).date()
        except ValueError:
            pass
    adm_logger.log_workflow(f'Invalid date or date format in analys_info', add=date_str, level=adm_logger.ERROR)
    adm_logger.log_workflow(adm_logger.feedback.invalid_date_in_analys_info(date_str), level=adm_logger.ERROR, purpose=adm_logger.FEEDBACK)
    return ''
