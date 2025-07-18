# -*- coding: utf-8 -*-

import datetime
import logging
import pathlib
from typing import Protocol

import openpyxl
import pandas as pd

from sharkadm.sharkadm_logger import adm_logger

logger = logging.getLogger(__name__)

DATE_FORMATS = ["%Y-%m-%d", "%Y-%m"]


class Mapper(Protocol):
    def get_internal_name(self, external_par: str) -> str: ...


class SamplingInfo:
    def __init__(self, data: dict, mapper: Mapper = None) -> None:
        self._data = data
        self._path = data.pop("path", None)
        self._mapper = mapper

        if self._mapper:
            self._map_data()

    def _map_data(self):
        if not self._mapper:
            adm_logger.log_workflow(
                f"No mapper found when trying to map AnalyseInfo file: {self._path}"
            )
            return
        new_data = {}
        for par, info_list in self._data.items():
            internal_par = self._mapper.get_internal_name(par)
            if internal_par.startswith("COPY_VARIABLE"):
                internal_par = internal_par.split(".")[1]
            new_data.setdefault(internal_par, [])
            for info in info_list:
                new_info = dict()
                for key, value in info.items():
                    internal_key = self._mapper.get_internal_name(key)
                    new_info[internal_key] = value
                new_data[internal_par].append(new_info)
        self._data = new_data

    def __str__(self):
        lst = "\n".join([sorted(self.data)])
        return f"Sampling info for parameters: \n{lst}"

    def get_info(self, par: str, date: datetime.date) -> dict:
        info = {}
        for info in self.data.get(par, []):
            if (
                info["VALIDFR"]
                and info["VALIDTO"]
                and (info["VALIDFR"] <= date <= info["VALIDTO"])
            ):
                return info
            if info["VALIDFR"] and (info["VALIDFR"] <= date):
                return info
            if info["VALIDTO"] and (date <= info["VALIDTO"]):
                return info
        return info

    def get(self, par: str, col: str, date: datetime.date) -> str | None:
        info = self.get_info(par=par, date=date)
        return info.get(col)

    @classmethod
    def from_txt_file(
        cls, path: str | pathlib.Path, mapper: Mapper = None, encoding: str = "cp1252"
    ) -> "SamplingInfo":
        path = pathlib.Path(path)
        if path.suffix != ".txt":
            msg = f"File is not a valid sampling_info text file: {path}"
            logger.error(msg)
            raise FileNotFoundError(msg)
        data = dict()
        data["path"] = path
        with open(path, encoding=encoding) as fid:
            header = []
            for r, line in enumerate(fid):
                if not line.strip():
                    continue
                split_line = [item.strip() for item in line.split("\t")]
                if r == 0:
                    header = split_line
                    continue
                if len(split_line) != len(header):
                    adm_logger.log_workflow(
                        f"No or insufficient data in sampling_info for parameter: "
                        f"{split_line[0]}",
                        level=adm_logger.WARNING,
                    )
                    continue
                line_dict = dict(zip(header, split_line))
                line_dict["VALIDFR"] = _get_date(line_dict["VALIDFR"])
                line_dict["VALIDTO"] = _get_date(line_dict["VALIDTO"])
                par = line_dict["PARAM"]
                data.setdefault(par, [])
                data[par].append(line_dict)
        return SamplingInfo(data, mapper=mapper)

    @classmethod
    def from_dv_template(
        cls, path: str | pathlib.Path, mapper: Mapper = None
    ) -> "SamplingInfo":
        wb = openpyxl.load_workbook(path)
        sheet_name = None
        for name in ["Provtagningsinfo"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            adm_logger.log_workflow(
                f"Could not find sampling_info sheet in file: {path}",
                level=adm_logger.WARNING,
            )
            return
            # raise Exception(f'Could not find analyse_info sheet in file: {path}')

        ws = wb.get_sheet_by_name(sheet_name)
        skip_nr_rows = 0
        for r in range(1, 5):
            if ws.cell(r, 1).value in ["Tabellhuvud:"]:
                skip_nr_rows = r - 1
                break
        df = pd.read_excel(path, skiprows=skip_nr_rows, sheet_name=sheet_name, dtype=str)
        df.fillna("", inplace=True)
        data = dict()
        data["path"] = path
        for row in df.iterrows():
            line_dict = row[1].to_dict()
            if not line_dict["PARAM"]:
                continue
            line_dict["VALIDFR"] = _get_date(line_dict["VALIDFR"])
            line_dict["VALIDTO"] = _get_date(line_dict["VALIDTO"])
            par = line_dict["PARAM"]
            data.setdefault(par, [])
            data[par].append(line_dict)
        return SamplingInfo(data, mapper=mapper)

    @property
    def data(self) -> dict[str, str]:
        return self._data

    @property
    def parameters(self) -> list[str]:
        """Returns a list of all parameters in the file. The list is unsorted."""
        return list(self._data)

    @property
    def columns(self) -> list[str]:
        return list(next(iter(self.data[next(iter(self._data))])))


def _get_date(date_str: str) -> datetime.date | str:
    if not date_str:
        return ""
    d_string = date_str.split()[0]
    for form in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(d_string, form).date()
        except ValueError:
            pass
    adm_logger.log_workflow(
        "Invalid date or date format in sampling_info",
        item=date_str,
        level=adm_logger.ERROR,
    )
    adm_logger.log_workflow(
        adm_logger.feedback.invalid_date_in_analys_info(date_str),
        level=adm_logger.ERROR,
        purpose=adm_logger.FEEDBACK,
    )
    return ""
