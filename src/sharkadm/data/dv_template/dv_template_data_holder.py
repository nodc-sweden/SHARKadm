# -*- coding: utf-8 -*-
import logging
import pathlib
import re

import openpyxl
import pandas as pd

from sharkadm import adm_logger
from sharkadm import config
# from sharkadm.config import get_data_type_mapper
from sharkadm.config.import_matrix import ImportMatrixConfig
from sharkadm.config.import_matrix import ImportMatrixMapper
from sharkadm.data import data_source
from sharkadm.data.archive import delivery_note
from sharkadm.data.archive import analyse_info
from sharkadm.data.archive import sampling_info
from sharkadm.data.data_holder import DataHolder

logger = logging.getLogger(__name__)


class DvTemplateDataHolder(DataHolder):
    _data_type: str | None = None
    _data_type_internal: str | None = None
    _data_format: str | None = None

    _date_str_format = '%Y-%m-%d'

    def __init__(self, template_path: str | pathlib.Path = None):
        super().__init__()
        self._template_path = pathlib.Path(template_path)
        adm_logger.dataset_name = self._template_path.name

        self._data: pd.DataFrame = pd.DataFrame()
        self._dataset_name: str | None = None

        self._mandatory_reg_columns: list = []
        self._mandatory_nat_columns: list = []

        self._delivery_note: delivery_note.DeliveryNote | None = None
        self._analyse_info: analyse_info.AnalyseInfo | None = None
        self._sampling_info: sampling_info.SamplingInfo | None = None

        self._import_matrix: ImportMatrixConfig | None = None
        self._import_matrix_mapper: ImportMatrixMapper | None = None
        # self._data_type_mapper = get_data_type_mapper()

        self._data_sources = {}

        self._initiate()
        self._load_delivery_note()
        self._load_mandatory_columns()
        self._load_import_matrix()
        self._load_data()
        self._load_delivery_note()  # Reload to map
        self._load_analyse_info()
        self._load_sampling_info()
        self._map_mandatory_lists()

    @staticmethod
    def get_data_holder_description() -> str:
        return """Holds data from a Data host delivery templates"""

    def _initiate(self) -> None:
        self._dataset_name = self.template_path.stem

    @property
    def received_data_files(self) -> list[pathlib.Path]:
        return [self._template_path]

    @property
    def header_mapper(self):
        return self._import_matrix_mapper

    @property
    def data_type(self) -> str:
        # return self._data_type_mapper.get(self.data_format)
        return self._data_type

    @property
    def data_type_internal(self) -> str:
        # return self._data_type_mapper.get(self.data_format)
        return self._data_type

    @property
    def dataset_name(self) -> str:
        return self._dataset_name

    @property
    def template_path(self) -> pathlib.Path:
        return self._template_path

    @property
    def delivery_note(self) -> delivery_note.DeliveryNote:
        return self._delivery_note

    @property
    def analyse_info(self) -> analyse_info.AnalyseInfo:
        return self._analyse_info

    @property
    def sampling_info(self) -> sampling_info.SamplingInfo:
        return self._sampling_info

    @property
    def import_matrix_mapper(self) -> ImportMatrixMapper:
        return self._import_matrix_mapper

    @property
    def mandatory_reg_columns(self):
        return self._mandatory_reg_columns

    @property
    def mandatory_nat_columns(self):
        return self._mandatory_nat_columns

    @property
    def min_year(self) -> str:
        return str(min(self.data['datetime']).year)

    @property
    def max_year(self) -> str:
        return str(max(self.data['datetime']).year)

    @property
    def min_date(self) -> str:
        return min(self.data['datetime']).strftime(self._date_str_format)

    @property
    def max_date(self) -> str:
        return max(self.data['datetime']).strftime(self._date_str_format)

    @property
    def min_longitude(self) -> str:
        return str(min(self.data['sample_longitude_dd'].astype(float)))

    @property
    def max_longitude(self) -> str:
        return str(max(self.data['sample_longitude_dd'].astype(float)))

    @property
    def min_latitude(self) -> str:
        return str(min(self.data['sample_latitude_dd'].astype(float)))

    @property
    def max_latitude(self) -> str:
        return str(max(self.data['sample_latitude_dd'].astype(float)))

    def _load_delivery_note(self) -> None:
        self._delivery_note = delivery_note.DeliveryNote.from_dv_template(self._template_path, mapper=self._import_matrix_mapper)

    def _load_analyse_info(self) -> None:
        self._analyse_info = analyse_info.AnalyseInfo.from_dv_template(self._template_path, mapper=self._import_matrix_mapper)

    def _load_sampling_info(self) -> None:
        self._sampling_info = sampling_info.SamplingInfo.from_dv_template(self._template_path, mapper=self._import_matrix_mapper)

    def _load_mandatory_columns(self):
        dn = pd.read_excel(self._template_path, sheet_name='Kolumnförklaring')
        key_col_name = dn.columns[3]
        dn['key_row'] = dn[key_col_name].apply(lambda x: True if type(x) == str and x.isupper() else False)

        fdn = dn[dn['key_row']]
        fdn = fdn.reset_index(drop=True)
        fdn['man_str'] = fdn[dn.columns[0]].apply(str)

        self._mandatory_reg_columns = list(fdn[fdn[dn.columns[0]] == '*'][key_col_name])
        # print(f'{dn.columns[0]=}')
        # print(f'{dn.columns=}')
        # print(fdn['man_str'])
        try:
            self._mandatory_nat_columns = list(fdn[fdn['man_str'].str.contains('*')][key_col_name])
        except re.error:
            adm_logger.log_workflow(f'Could not find mandatory_nat_columns', level=adm_logger.WARNING)

    def _map_mandatory_lists(self):
        if not self.import_matrix_mapper:
            adm_logger.log_workflow(f'Could not map mandatory lists in {self.__class__.__name__}')
            return
        self._mandatory_reg_columns = [self._import_matrix_mapper.get_internal_name(col) for col in
                                       self._mandatory_reg_columns]
        self._mandatory_nat_columns = [self._import_matrix_mapper.get_internal_name(col) for col in
                                       self._mandatory_nat_columns]

    def _load_import_matrix(self) -> None:
        """Loads the import matrix for the given data type and provider found in delivery note"""
        data_type_mapper = config.get_data_type_mapper()
        dtype = data_type_mapper.get(self.delivery_note.data_type, default=self.delivery_note.data_type.lower())
        self._import_matrix = config.get_import_matrix_config(data_type=dtype)
        if not self._import_matrix:
            msg = f'Could not find import matrix for data_type: {self.delivery_note.data_type}'
            adm_logger.log_workflow(msg)
            return
        try:
            self._import_matrix_mapper = self._import_matrix.get_mapper(self.delivery_note.import_matrix_key)
        except KeyError as e:
            msg = f'Could not get import matrix mapper for import_matrix_key: {self.delivery_note.import_matrix_key}'
            adm_logger.log_workflow(msg)

    def _add_concatenated_column(self, new_column: str, columns_to_use: list[str]) -> None:
        """Adds a concatenated column specified in new_column using the columns listed in columns_to_use"""
        for col in columns_to_use:
            if new_column not in self._data.columns:
                self._data[new_column] = self._data[col]
            else:
                self._data[new_column] = self._data[new_column] + self._data[col]

    def _check_data_source(self, data_source: data_source.DataFile) -> None:
        #if self._data_type_mapper.get(data_source.data_type) != self._data_type_mapper.get(self.data_type):
        if data_source.data_type.lower() != self.data_type.lower():
            msg = f'Data source {data_source} is not of type {self.data_type}'
            logger.error(msg)
            raise ValueError(msg)

    @staticmethod
    def _get_data_from_data_source(data_source: data_source.DataFile) -> pd.DataFrame:
        data = data_source.get_data()
        data = data.fillna('')
        data.reset_index(inplace=True, drop=True)
        return data

    # def _concat_data_source(self, data_source: data_source.DataFile) -> None:
    #     self._check_data_source(data_source)
    #     self._data_sources[str(data_source)] = data_source
    #     """Concats new data source to self._data"""
    #     new_data = self._get_data_from_data_source(data_source)
    #     new_data['data_source'] = data_source.source
    #     new_data['dataset_name'] = self._dataset_name
    #     self._data = pd.concat([self._data, new_data])
    #     self._data.fillna('', inplace=True)
    #     self._data.reset_index(inplace=True, drop=True)

    def _set_data_source(self, data_source: data_source.DataFile) -> None:
        """Sets a single data source to self._data"""
        self._data_type = data_source.data_type
        self._add_data_source(data_source)
        self._data = self._get_data_from_data_source(data_source)

    def _add_data_source(self, data_source: data_source.DataFile) -> None:
        """Adds a data source to instance variable self._data_sources. This method is not adding to data itself."""
        self._check_data_source(data_source)
        self._data_sources[str(data_source)] = data_source

    def _load_data(self):
        wb = openpyxl.load_workbook(self._template_path)
        sheet_name = None
        for name in ['data', 'Klistra in i denna', 'Klistra in  i denna', 'Kolumner']:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            raise Exception(f'Could not find data sheet in file: {self._template_path}')

        ws = wb.get_sheet_by_name(sheet_name)
        for r in range(1, 5):
            if ws.cell(r, 1).value in ['Tabellhuvud:']:
                self._number_metadata_rows = r - 1
                break
        d_source = data_source.XlsxFormatDataFile(path=self._template_path, data_type=self.delivery_note.data_type,
                                                  sheet_name=sheet_name, skip_rows=self.number_metadata_rows)
        if self.import_matrix_mapper:
            d_source.map_header(self.import_matrix_mapper)

        self._set_data_source(d_source)

